#!/usr/bin/env python3
"""
scripts/seed_catalog.py — заливка каталога подшипников из XLSX в D1.

Использование:
    python3 scripts/seed_catalog.py --xlsx catalog.xlsx --d1 b24-catalog-db [--remote]

Ожидаемые колонки XLSX (маппинг в CATALOG_FIELDS, можно переопределить --mapping):
    part_number, brand, family, name, d, D, H, mass,
    execution, clearance, class, analog_gost, analog_iso,
    price, currency, stock, image_url

Скрипт генерирует SQL-файл /tmp/catalog_seed.sql и вызывает wrangler d1 execute.
"""

import argparse
import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Требуется openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

CATALOG_FIELDS = [
    "part_number", "brand", "family", "name",
    "d", "D", "H", "mass",
    "execution", "clearance", "class",
    "analog_gost", "analog_iso",
    "price", "currency", "stock", "image_url",
]


def sql_quote(v):
    if v is None or v == "":
        return "NULL"
    if isinstance(v, (int, float)):
        return str(v)
    s = str(v).replace("'", "''")
    return f"'{s}'"


def build_sql(rows):
    lines = ["BEGIN;"]
    cols = ", ".join(CATALOG_FIELDS + ["updated_at"])
    placeholders = "strftime('%s','now') * 1000"
    for r in rows:
        values = [sql_quote(r.get(f)) for f in CATALOG_FIELDS] + [placeholders]
        lines.append(
            f"INSERT OR REPLACE INTO catalog ({cols}) VALUES ({', '.join(values)});"
        )
    lines.append("COMMIT;")
    return "\n".join(lines)


def parse_xlsx(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    headers = [
        (c.value or "").strip().lower() if isinstance(c.value, str) else c.value
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        rec = {}
        for h, v in zip(headers, row):
            if h in CATALOG_FIELDS:
                rec[h] = v
        if rec.get("part_number"):
            rows.append(rec)
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Путь к XLSX каталогу")
    ap.add_argument("--d1", default="b24-catalog-db", help="Имя D1 БД")
    ap.add_argument("--remote", action="store_true", help="Применять на remote D1 (прод)")
    ap.add_argument("--out", default="/tmp/catalog_seed.sql", help="SQL-файл")
    ap.add_argument("--dry-run", action="store_true", help="Только сгенерировать SQL, не применять")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        print(f"Файл не найден: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    rows = parse_xlsx(xlsx_path)
    print(f"Прочитано строк: {len(rows)}")
    if not rows:
        sys.exit(0)

    sql = build_sql(rows)
    Path(args.out).write_text(sql, encoding="utf-8")
    print(f"SQL записан: {args.out} ({len(sql)} байт)")

    if args.dry_run:
        print("DRY-RUN — применение пропущено")
        return

    cmd = ["npx", "wrangler", "d1", "execute", args.d1, f"--file={args.out}"]
    if args.remote:
        cmd.append("--remote")
    print("Выполняю:", " ".join(cmd))
    result = subprocess.run(cmd, check=False)
    sys.exit(result.returncode)


if __name__ == "__main__":
    main()
